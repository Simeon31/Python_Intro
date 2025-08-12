import openpyxl

# wb = openpyxl.Workbook() // Creates a new instance

wb = openpyxl.load_workbook("transactions.xlsx") # loads existing file
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["A1"]
cells = sheet["A:C"] # accessing with range
print(cell.value)
print(cells)

sheet.append([1, 2, 9]) # adds row at the end of the sheet
# sheet.insert_rows/cols() # inserts a row at a given index
# sheet.delete_rows/cols()

print("Column: ", cell.column)
print("Row: ", cell.row)
print("Coordinate: ", cell.coordinate)

# Iteration
# Note: The command query separation principle should be taken into account

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=column) # Another way of accessing a cell
        print(cell.value)

wb.save("transactions2.xlsx")